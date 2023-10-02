import os
import tkinter as tk
from tkinter import messagebox
from ADmin import addproduct
from ADmin import Fournisseur
from ADmin import Type
from ADmin import Findproduct
from ADmin import LieuService
from ADmin import Service
from ADmin import supA
import mysql.connector

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

conn = mysql.connector.connect(
    host="localhost",
    user="root",
    port='3306',
    database="fmpdf"
)
cursor = conn.cursor()

def adminpanel():
    def validate_credentials():
        username = entry_name.get()
        password = entry_passwd.get()

        # Vérification des informations d'identification dans la base de données
        query = "SELECT * FROM admin WHERE name=%s AND passwd=%s"
        cursor.execute(query, (username, password))
        result = cursor.fetchone()

        if result:
            def ADDproduct():
                addproduct.Add_pro()

            def ADDfournisseur():
                Fournisseur.Add_fournisseur()

            def ADDtype():
                Type.ADD_type()

            def ADDservice():
                Service.ADD_services()

            def Addlieuservice():
                LieuService.ADD_lieu()

            def Findproducts():
                Findproduct.Find_pro()

            def Export_Etat_stock():
                # Fetch data from the 'equipement' table
                query = "SELECT * FROM equipement"
                cursor.execute(query)
                results = cursor.fetchall()

                # Create a pandas DataFrame from the fetched data
                column_names = [col[0] for col in cursor.description]
                df = pd.DataFrame(results, columns=column_names)

                # Export the DataFrame to an Excel file
                output_file = "equipement_stock.xlsx"
                df.to_excel(output_file, index=False)

                # Apply formatting to the Excel file
                wb = openpyxl.load_workbook(output_file)
                ws = wb.active

                # Set column width to 150 pixels
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[col[0].column_letter].width = adjusted_width if adjusted_width > 20 else 20

                # Set header row background color to yellow
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # Save the modified Excel file
                try:
                    wb.save(output_file)
                except PermissionError:
                    os.remove(output_file)
                    wb.save(output_file)

                # Show a message to inform the user about the successful export
                messagebox.showinfo("Export Success", f"Data exported to '{output_file}'")
            def SUPP():
                supA.SUPP_pro()

            fenetre_Admin.destroy()
            fenetre_Gestion_admin =tk.Toplevel()
            fenetre_Gestion_admin.title("Admin Gestion")

            largeur = 500
            hauteur = 400
            fenetre_Gestion_admin.geometry(f"{largeur}x{hauteur}")
            icone_path = 'C:\\Users\\WD\\Desktop\\stage\\Prjt\\ADmin\\imgs\\nose_120696.ico'

            # Charger l'icône personnalisée
            fenetre_Gestion_admin.iconbitmap(icone_path)

            image_path9 = 'C:\\Users\\WD\\Desktop\\stage\\Prjt\\ADmin\\imgs\\Dash.png'
            background_image9 = tk.PhotoImage(file=image_path9)

            canvas2 = tk.Canvas(fenetre_Gestion_admin, width=largeur, height=hauteur)
            canvas2.pack()
            canvas2.create_image(0, 0, image=background_image9, anchor=tk.NW)

            btn_ADDproduct_coords = (140, 180)
            btn_ADDfournisseur_coords = (140, 230)
            btn_ADDtype_coords = (140, 280)
            btn_ADDservice_coords = (140, 330)

            btn_Addlieuservice_coords = (350, 180)
            btn_Findproduct_coords = (350, 230)
            btn_Export_Etat_stock_coords = (350, 280)
            btn_Export_bon_appo_coords = (350, 330)

            btn_ADDproduct = tk.Button(fenetre_Gestion_admin, text="Ajouter un Article", bg="darkslategray",fg="white", width=20,command=ADDproduct)
            btn_ADDfournisseur = tk.Button(fenetre_Gestion_admin, text="Ajouter un foutnisseur", bg="darkslategray", fg="white", width=20, command=ADDfournisseur)
            btn_ADDtype = tk.Button(fenetre_Gestion_admin, text="Ajouter Type Article", bg="darkslategray", fg="white", width=20, command=ADDtype)
            btn_ADDservice = tk.Button(fenetre_Gestion_admin, text="Informations sur Article", bg="darkslategray", fg="white", width=20, command=Findproducts)

            btn_Addlieuservice = tk.Button(fenetre_Gestion_admin, text="Ajouter service", bg="darkslategray", fg="white", width=20,command=ADDservice)
            btn_Findproduct = tk.Button(fenetre_Gestion_admin, text="Ajouter lieu service", bg="darkslategray", fg="white", width=20,command=Addlieuservice)
            btn_Export_Etat_stock = tk.Button(fenetre_Gestion_admin, text="Export Etat stock", bg="darkslategray", fg="white", width=20,command= Export_Etat_stock)
            btn_Export_bon_appo = tk.Button(fenetre_Gestion_admin, text="Supprimer Article", bg="darkslategray", fg="white", width=20,command=SUPP)


            btn_ADDproduct_window = canvas2.create_window(btn_ADDproduct_coords[0], btn_ADDproduct_coords[1],window=btn_ADDproduct)
            btn_ADDfournisseur_window = canvas2.create_window(btn_ADDfournisseur_coords[0],btn_ADDfournisseur_coords[1],window=btn_ADDfournisseur)
            btn_ADDtype_window = canvas2.create_window(btn_ADDtype_coords[0],btn_ADDtype_coords[1], window=btn_ADDtype)
            btn_ADDservice_window = canvas2.create_window(btn_ADDservice_coords[0],btn_ADDservice_coords[1], window=btn_ADDservice)

            btn_ADDproduct_window = canvas2.create_window(btn_Addlieuservice_coords[0], btn_Addlieuservice_coords[1],window=btn_Addlieuservice)
            btn_ADDfournisseur_window = canvas2.create_window(btn_Findproduct_coords[0],btn_Findproduct_coords[1],window=btn_Findproduct)
            btn_ADDtype_window = canvas2.create_window(btn_Export_Etat_stock_coords[0],btn_Export_Etat_stock_coords[1], window=btn_Export_Etat_stock)
            btn_ADDservice_window = canvas2.create_window(btn_Export_bon_appo_coords[0],btn_Export_bon_appo_coords[1], window=btn_Export_bon_appo)

            fenetre_Gestion_admin.mainloop()

        else:
        # Afficher une boîte de message en cas d'erreur d'identification
         messagebox.showerror("Erreur", "Nom d'utilisateur ou mot de passe incorrect")


    fenetre_Admin = tk.Toplevel()
    fenetre_Admin.title("ADMIN")

    largeur = 500
    hauteur = 400
    fenetre_Admin.geometry(f"{largeur}x{hauteur}")
    image_path1 = 'C:\\Users\\WD\\Desktop\\stage\\Prjt\\ADmin\\imgs\\SECOND.png'
    background_image1 = tk.PhotoImage(file=image_path1)

    icone_path = 'C:\\Users\\WD\\Desktop\\stage\\Prjt\\ADmin\\imgs\\nose_120696.ico'
    # Charger l'icône personnalisée
    fenetre_Admin.iconbitmap(icone_path)

    # Création d'un canevas pour afficher l'image en tant que fond d'écran
    canvas1 = tk.Canvas(fenetre_Admin, width=largeur, height=hauteur)
    canvas1.pack()
    canvas1.create_image(0, 0, image=background_image1, anchor=tk.NW)
    entry_name = tk.Entry(fenetre_Admin)
    entry_passwd = tk.Entry(fenetre_Admin, show="*")
    # Coordonnées des boutons sur l'image de fond
    btn_Valider_coords= (364, 310)
    btn_Valider = tk.Button(canvas1, text="ENTREZ", bg="Green",fg="white", width=8 , command=validate_credentials)
    entry_name.place(relx=0.376, rely=0.667, anchor=tk.W)
    entry_passwd.place(relx=0.376, rely=0.775, anchor=tk.W)
    # Placement des boutons et des autres widgets sur l'image de fond
    canvas1.create_window(btn_Valider_coords[0], btn_Valider_coords[1], window=btn_Valider)
    fenetre_Admin.mainloop()
